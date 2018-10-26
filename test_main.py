import pytest
from main import plot_data
import os
import pandas as pd
from main import calc_duration
from main import calc_v_extreme
from pytest import approx
from main import user_input
from main import is_data_valid
from main import edge_case
from main import check_spacing
from main import Hilbert
import numpy.fft as fft
from main import create_metrics
from main import write_json
from main import write_excel
from openpyxl import load_workbook
from main import calc_avg
from main import peak_detector
from main import threshold_peak_detect
from main import check_loop
import numpy as np


@pytest.mark.parametrize("file, expected", [
    ('sine.csv', (1.0, -1.0)),
]
                         )
def test_calc_v_extreme(file, expected):
    """

    Args:
        file: file name (string)
        expected: max and min of voltage data

    Returns: Pass or Fail

    """
    headers = ['time', 'voltage']
    data = pd.read_csv(file, names=headers)
    store = calc_v_extreme(data)
    store = (round(store[0]), round(store[1]))
    assert expected == store


@pytest.mark.parametrize("file, expected", [
    ('sine.csv', 11.85)
])
def test_calc_duration(file, expected):
    """

    Args:
        file: file name
        expected: length of time data

    Returns: Pass or Fail

    """
    headers = ['time', 'voltage']
    data = pd.read_csv(file, names=headers)
    dur = calc_duration(data)
    assert expected == dur


@pytest.mark.parametrize("file, dur, window, expected", [
    ('sine.csv', 11.9, [], list([11.9, False])),
    ('sine.csv', 11.9, [-1, 11.9], list([11.9, False])),
    ('sine.csv', 11.9, [1], list([11.9, False])),
    ('sine.csv', 11.9, ['dog', 1], list([11.9, False])),
])
def test_user_input(file, dur, window, expected):
    """

    Args:
        file: file name
        dur: length of time data
        window: user input window to be tested
        expected: functions response to user input

    Returns: Pass or Fail

    """
    out = user_input(dur, window)
    assert expected == out


@pytest.mark.parametrize("file, expected", [
    ('sine.csv', True),
    ('sine_with_words.csv', False),
])
def test_is_data_valid(file, expected):
    """

    Args:
        file: file name
        expected: true if all values could be case to a float;
        false if strings were present

    Returns: Pass or Fail

    """
    headers = ['time', 'voltage']
    data = pd.read_csv(file, names=headers)
    valid_data = is_data_valid(data)
    switch = True
    for x in valid_data['time']:
        y = float(x)
        if x != y:
            switch = False
    for x in valid_data['voltage']:
        y = float(x)
        if x != y:
            switch = False
    assert expected == switch


@pytest.mark.parametrize("file, expected", [
    ('sine.csv', (0.017, 0.017)),
])
def test_edge_case(file, expected):
    """

    Args:
        file: file name
        expected:

    Returns: Pass or Fail

    """
    headers = ['time', 'voltage']
    data = pd.read_csv(file, names=headers)
    avg_v = 0
    for x in data['voltage']:
        avg_v += float(x)
    avg_v /= len(data['voltage'])
    out = edge_case(data, 200, avg_v)
    tup = (out.loc[0]['voltage'], out.loc[len(out)-1]['voltage'])
    tup_out = (round(tup[0], 3), round(tup[1], 3))

    assert expected == tup_out


@pytest.mark.parametrize("file, expected, space, one, two", [
    ('sine.csv', False, 1, 10, 12),
    ('sine.csv', True, 0.5, 1, 120),
])
def test_check_spacing(file, expected, space, one, two):
    """

    Args:
        file: name of file
        expected: True if appropriately spaced, false if not
        space: the distance the function is ensuring between detected peaks
        one: index
        two: index

    Returns: Pass or Fail

    """
    headers = ['time', 'voltage']
    data = pd.read_csv(file, names=headers)
    found = list()
    found.append([one, data.loc[one]['time'], data.loc[one]['voltage']])
    found.append([two, data.loc[two]['time'], data.loc[two]['voltage']])
    headers = ['index', 'time', 'voltage']
    return_df = pd.DataFrame(found, columns=headers)
    out = check_spacing(return_df, data, space)
    assert expected == out


@pytest.mark.parametrize("file, expected", [
    ('sine.csv', True)
])
def test_Hilbert(file, expected):
    """

    Args:
        file: file name
        expected:

    Returns: Pass or Fail

    """
    headers = ['time', 'voltage']
    data = pd.read_csv(file, names=headers)
    data_lpf = Hilbert(data, 0.0001)
    if not np.isnan(np.sum(data_lpf)):
        out = True
    else:
        out = False
    assert expected == out


@pytest.mark.parametrize("file, expected", [
    ('sine.csv', type({}))
])
def test_create_metrics(file, expected):
    """

    Args:
        file: file name
        expected: type < class dict>

    Returns: Pass or Fail

    """
    headers = ['time', 'voltage']
    data = pd.read_csv(file, names=headers)
    out = create_metrics(data, 1, 1, 1)
    assert expected == type(out)


@pytest.mark.parametrize("file, expected", [
    ('sine.csv', True)
])
def test_write_json(file, expected):
    """

    Args:
        file: file name
        expected: True if json file can be found;
        False if file cannot be found

    Returns: Pass or Fail

    """
    metrics = {'thing_one': 1, 'thing_two': 2}
    write_json(file, metrics)
    json_file_name = file.split('.')[0]
    out = False
    for file in os.listdir(os.getcwd()):
        if file == json_file_name + '.json':
            out = True
    assert expected == out


@pytest.mark.parametrize("file, expected", [
    (list([1, 1, 1, 1, 1, 1, 1, 1, 1, 1]),
     list([1, 1, 1, 1, 1, 1, 1, 1, 1, 1]))
])
def test_write_excel(file, expected):
    """

    Args:
        file: file name
        expected: values written to excel

    Returns: Pass or Fail

    """
    write_excel(file, expected, 'test_write_excel.xlsx')
    wb = load_workbook('test_write_excel.xlsx')
    ws = wb.active
    out = int(ws['C' + str(int(file[0]) + 1)].value)
    assert expected[0] == out


@pytest.mark.parametrize("interval, found, dur, expected", [
    ([10, False], [1, 2, 3, 4, 5], 5, 60),
    ([(1, 3), True], [1, 2, 3, 4, 5], 5, 30),
])
def test_calc_avg(interval, found, dur, expected):
    """

    Args:
        interval: User chosen interval
        found: data frame containing index, time, and voltage of found peaks
        dur: length of time framed by original data
        expected: BPM in given interval

    Returns: Pass or Fail

    """
    headers = ['time']
    found = pd.DataFrame(found, columns=headers)
    bpm = calc_avg(interval, found, dur)
    bpm = int(bpm)
    assert expected == bpm


@pytest.mark.parametrize("file, expected", [
    ('test_data22.csv', 34),
])
def test_peak_detector(file, expected):
    """

    Args:
        file: file name
        expected: number of peaks in data

    Returns: Pass or Fail

    """
    headers = ['time', 'voltage']
    data = pd.read_csv(file, names=headers)
    filtered = Hilbert(data, 0.005)
    df = peak_detector(filtered, data)
    out = len(df['index'])
    assert expected == out


@pytest.mark.parametrize("file, expected, sign", [
    ('sine.csv', 2, 1),
    ('sine.csv', 2, -1),
])
def test_threshold_peak_detect(file, expected, sign):
    """

    Args:
        file: file name
        expected: number of peaks according to associated threshold
        sign: detector creates positive or negative threshold

    Returns: Pass or Fail

    """
    headers = ['time', 'voltage']
    data = pd.read_csv(file, names=headers)
    found = threshold_peak_detect(data, (data['voltage'].max(),
                                         data['voltage'].min()), sign)
    out = len(found['index'])
    assert expected == out


@pytest.mark.parametrize("file, index, method, expected, user_interval", [
    ('sine.csv', [32, 156], 1, True, [0, -1]),
    ('sine.csv', [32, 156], 0, True, [0, -1]),
])
def test_plot_data(file, index, method, expected, user_interval):
    """

    Args:
        file: file name
        index: indices of peaks found
        method: which type of peak detection method used
        expected: if plot was successful True
        user_interval: interval put in by the user
        before it is checked for validity

    Returns: Pass or Fail

    """
    headers = ['time', 'voltage']
    data = pd.read_csv(file, names=headers)
    plot_data(data, data['voltage'], index, file, method, user_interval)
    switch = True
    assert switch == expected


@pytest.mark.parametrize("filter_value, file, space, print_plot, "
                         "expected, found_state",
                         [(0.005, 'sine.csv', 50, 0, 2, 0),
                          (0.005, 'sine.csv', 50, 0, 2, 1)
                          ])
def test_check_loop(filter_value, file, space, print_plot,
                    expected, found_state):
    """

    Args:
        filter_value: Cutoff frequency for butterworth filter
        file: file name
        space: distance ensured between peaks found
        print_plot: 1 for show plots; 0 for don't show
        expected: number of peaks expected
        found_state: tests empty and populated found data frame

    Returns: Pass or Fail

    """
    headers = ['time', 'voltage']
    data = pd.read_csv(file, names=headers)
    max_min = (data['voltage'].max(), data['voltage'].min())
    if found_state:
        found = [32, data.loc[32]['time'], data.loc[32]['voltage']]
        found = found.append([156, data.loc[156]['time'],
                              data.loc[156]['voltage']])
    else:
        found = []
    headers = ['index', 'time', 'voltage']
    found = pd.DataFrame(found, columns=headers)
    found2 = check_loop(found, data, filter_value, file,
                        space, print_plot, max_min)
    assert expected == len(found2['index'])
