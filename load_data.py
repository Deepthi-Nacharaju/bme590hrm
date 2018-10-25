import pandas as pd
import os
import matplotlib.pyplot as plt
import numpy as np
import sys
import scipy.signal as signal
from scipy.signal import hilbert
import json
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def plot_data(data, filtered, index, file):
    """

    Args:
        data: input data with padded ends
        filtered: low pass enveloped data
        index: locations of found peaks
        file: name of csv file

    Returns:

    """

    data_points = []
    headers = ['time', 'voltage']
    for x in index:
        data_points.append([data.loc[x]['time'], filtered[x]])

    data_points_df = pd.DataFrame(data_points, columns=headers)
    plt.plot(data['time'], data['voltage'])
    plt.scatter(data_points_df['time'], data_points_df['voltage'], c='red')
    plt.plot(data['time'], filtered)
    plt.axis('tight')
    plt.ylabel('Voltage')
    plt.xlabel('Time (s)')
    plt.title('ECG with Peak Detection: ' + str(file))
    plt.legend(['ECG', 'LPF Envelope', 'Detected Peak'])
    plt.show()
    logging.info('This only outputs a plot')
    return


def calc_duration(data):
    """

    Args:
        data: input data before padding

    Returns:
        dur: duration of input data

    """
    try:
        dur = data.loc[data.index[-1]]['time']-data.loc[1]['time']
    except TypeError:
        dur = float(data.loc[data.index[-1]]['time']) - \
              float(data.loc[1]['time'])
    return dur


def calc_v_extreme(data):
    """

    Args:
        data: input data before padding

    Returns:
        store: tuple containing minimum and maximum voltage

    """
    max_val = data['voltage'].max()
    min_val = data['voltage'].min()
    store = (max_val, min_val)
    return store


def peak_detector(dx, dy, data):
    """

    Args:
        dx: adjusted time vector after differentiating voltage vector
        (off by one error)
        dy: differentiated data
        data: original data from csv file

    Returns:
        return_df: data frame containing index, time, and
        voltage point of isolated peaks
    """
    return_values = []
    y_old = dy[0]
    index_old = -999
    indices = []
    switch = False
    go = False
    for index, y in enumerate(dy):
        if y - y_old > 0 and float(data.loc[index]['time']) > 0:
            go = True
        if y - y_old <= 0 and index - index_old > 5 \
                and switch is False and go is True:
            if index_old == -999 or go is True:
                return_values.append([index, dx.loc[index]['time'], y])
                indices.append(index)
                index_old = index
                switch = True
        if y - y_old > 0 and go is True:
            switch = False
        y_old = y
    headers = ['index', 'time', 'voltage']
    return_df = pd.DataFrame(return_values, columns=headers)
    return return_df


def user_input(duration, window=None):
    """

    User Input for choosing window over which to take average
    Args:
        window: Optional input to determine window from within
                load_data.py
        duration: time duration of original data file

    Returns:
        out: user identified duration possibly by default
    """
    if not window:
        try:
            interval_one = sys.argv[1]
            interval_two = sys.argv[2]
            interval = list([float(interval_one), float(interval_two)])
            out = list([interval, True])
        except IndexError:
            interval = duration
            print('No Time Window Indicated. Default = ' + str(interval))
            out = list([interval, False])
        logging.debug('What type of error does this cause? --> IndexError')
        if interval > duration:
            print('User Input Exceeds Data Duration. Default = ' + str(duration))
            out = list([duration, False])
        return out
    else:
        try:
            interval_one = window[0]
            interval_two = window[1]
            interval = list([float(interval_one), float(interval_two)])
            out = list([interval, True])
        except TypeError:
            print('User input for window must be a tuple with two numbers')
            print('Default = ' + str(interval))
            out = list([interval, False])


def calc_avg(interval, found, dur):
    """

    Args:
        interval: determined interval from user_input
        found: data frame containing index, time,
        and voltage points of detected peaks
        dur: duration of data file

    Returns:
        bpm: the number of bpms as detected in the chosen duration
    """
    if not interval[1]:
        bpm = int(float(len(found['time']))/dur*60)
    else:
        bpm_range = interval[0]
        bpm_count = 0
        for x in found['time']:
            if x > bpm_range[0] and x < bpm_range[1]:
                bpm_count += 1
        bpm = float(bpm_count)/(float(bpm_range[1])-float(bpm_range[0]))*60

    return bpm


def create_metrics(found, extreme, dur, bpm):
    """

    Args:
        found: data frame containing index,
        time, and voltage points of found peaks
        extreme: max and min voltage found in data file
        dur: the time length of the data file
        bpm: number of beats per min in time interval requested by user

    Returns:
        metrics: dictionary requested by assignment
    """
    metrics = dict()
    metrics['voltage_extremes'] = extreme
    metrics['duration'] = dur
    metrics['num_beats'] = len(found['time'])
    metrics['mean_hr_bpm'] = bpm
    metrics['beats'] = list(found['time'])
    print(metrics)
    logging.info('Final Dictionary Creation')
    return metrics


def write_json(file, metrics):
    """

    Args:
        file: name of file
        metrics: dictionary containing necessary parameters for assignment

    Returns:
        saves a json file with the requested dictionary
    """
    json_name = file.split('.')[0] + '.json'
    with open(json_name, 'w') as outfile:
        json.dump(metrics, outfile)

    logging.info('make sure everything in metrics is a '
                 'dictionary and NOT a data frame')


def Hilbert(data, cutoff):
    """

    Args:
        data: input data with padding
        cutoff:

    Returns:
        filtered: low pass filtered enveloped data
    """
    analytic_signal = hilbert(data['voltage'])
    amplitude_envelope = np.abs(analytic_signal)
    n = 2  # Filter order
    wn = cutoff  # Cutoff frequency
    b, a = signal.butter(n, wn, output='ba')
    filtered = signal.filtfilt(b, a, amplitude_envelope)
    return filtered


def edge_case(data):
    """

    Args:
        data: raw input data from csv file

    Returns:
        data: raw input data with padding
    """
    extra = []
    headers = ['time', 'voltage']
    try:
        dt = data.loc[50]['time']-data.loc[49]['time']
    except TypeError:
        dt = float(data.loc[50]['time']) - float(data.loc[49]['time'])
    for x in range(0, 200):
        extra.append([dt*x+float(data.loc[len(data['time'])-1]['time']),
                      -0.25])
    extra = pd.DataFrame(extra, columns=headers)
    data = data.append(extra, sort=False)
    data = data.reset_index()
    extra2 = []
    for x in range(0, 200):
        extra2.append([float(data.loc[0]['time']) - dt*200 + dt * x, -0.25])
    extra2 = pd.DataFrame(extra2, columns=headers)
    data = extra2.append(data, sort=False)
    data = data.reset_index()
    return data


def check_spacing(found, data, space):
    """

    Args:
        space: ensure distance between found peaks
        is not more than this in seconds
        found: data frame containing index,
        time, and voltage points of found peaks
        data: imported data that has been padded

    Returns:
        boolean if peaks are too far apart = True
    """
    difference = []
    old_x = 0
    for x in found['index']:
        difference.append(data.loc[x]['time']-data.loc[old_x]['time'])
        old_x = x
    if max(difference) - sum(difference)/len(difference) > space:
        return True
    else:
        return False


def is_data_valid(data):
    """

    Args:
        data: raw input data frame

    Returns:
        data: data that has been cast to floats
    """
    dropped = 0
    for index_, y in enumerate(data['time']):
        try:
            data['time'][index_] = float(y)
        except ValueError:
            data = data.drop(data.index[index_])
            dropped += 1
    for index_, y in enumerate(data['voltage']):
        try:
            data['voltage'][index_] = float(y)
        except ValueError:
            data = data.drop(data.index[index_])
            dropped += 1
    return data


def check_loop(found, data, filter_value, file, space):
    """

    Args:
        space: ensuring this distance between peaks
        found: data frame containing index, time,
        and voltage of found peaks
        data: data that has been padded
        filter_value: cutoff frequency for butterworth filter
        file: name of csv file

    Returns:
        found: optimized data frame containing index,
        time, and voltage of found peaks
    """
    if not found.empty:
        check = check_spacing(found, data, space)
        counter = 0
        while check:
            filter_value += 0.002
            filtered = Hilbert(data, filter_value)
            plt.plot(data['time'], data['voltage'])
            plt.plot(data['time'], filtered)
            plt.title(str(file) + ' ' + str(counter))
            plt.show()
            dx = data.drop([0, 0])
            found = peak_detector(dx, filtered, data)
            check = check_spacing(found, data, 1)
            if counter == 3:
                check = False
            counter += 1

    return found


def write_excel(file_number, export_excel, excel_file_name):
    wb = load_workbook(excel_file_name)
    ws = wb.active
    counter = 0
    yellow_fill = PatternFill(start_color='FFFFFF00',
                              end_color='FFFFFF00',
                              fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFFFF',
                             end_color='FFFFFFFF',
                             fill_type='solid')
    orange_fill = PatternFill(start_color='FFFF8C00',
                              end_color='FFFF8C00',
                              fill_type='solid')
    red_fill = PatternFill(start_color='FFFF0000',
                           end_color='FFFF0000',
                           fill_type='solid')
    green_fill = PatternFill(start_color='FF00FF00',
                             end_color='FF00FF00',
                             fill_type='solid')
    grey_fill = PatternFill(start_color='FFC0C0C0',
                            end_color='FFC0C0C0',
                            fill_type='solid')
    for x in file_number:
        ws['C' + str(int(x) + 1)] = str(export_excel[counter])
        counter += 1
        try:
            if np.abs(int(ws['C' + str(int(x) + 1)].value) -
                      int(ws['B' + str(int(x) + 1)].value)) > 10:
                ws['C' + str(int(x) + 1)].fill = grey_fill
                ws['C' + str(int(x) + 1)] = str('Bad Data')
            elif np.abs(int(ws['C' + str(int(x) + 1)].value) -
                        int(ws['B' + str(int(x) + 1)].value)) > 5:
                ws['C' + str(int(x) + 1)].fill = red_fill
            elif np.abs(int(ws['C' + str(int(x) + 1)].value) -
                        int(ws['B' + str(int(x) + 1)].value)) > 2:
                ws['C' + str(int(x) + 1)].fill = orange_fill
            elif np.abs(int(ws['C' + str(int(x) + 1)].value) -
                        int(ws['B' + str(int(x) + 1)].value)) > 0:
                ws['C' + str(int(x) + 1)].fill = yellow_fill
            elif np.abs(int(ws['C' + str(int(x) + 1)].value) -
                        int(ws['B' + str(int(x) + 1)].value)) == 0:
                ws['C' + str(int(x) + 1)].fill = green_fill
            else:
                ws['C' + str(int(x) + 1)].fill = white_fill
        except TypeError:
            print('File ' + str(x) + ' has not been tracked')
            ws['C' + str(int(x) + 1)].fill = white_fill
    wb.save(excel_file_name)


def main():
    """

    Returns: saved json file and lots of plots of viewing

    """
    plt.close('all')
    new_path = os.getcwd() + '/data'
    os.chdir(new_path)
    headers = ['time', 'voltage']
    export_excel = list()
    file_number = list()
    space = 1
    for file in os.listdir(os.getcwd()):
        print(file)
        try:
            if file.split('.')[1] == 'csv':
                data = pd.read_csv(file, names=headers)
                extreme = calc_v_extreme(data)
                dur = calc_duration(data)
                interval = user_input(dur, (5, 7))
                data = is_data_valid(data)
                try:
                    dx = data.loc[3]['time'] - data.loc[2]['time']
                except TypeError:
                    dx = float(data.loc[3]['time']) - \
                         float(data.loc[2]['time'])
                data = edge_case(data)
                filter_value = 0.005
                filtered = Hilbert(data, filter_value)
                dx = data.drop([0, 0])
                found = peak_detector(dx, filtered, data)
                found = check_loop(found, data, filter_value, file, space)
                bpm = calc_avg(interval, found, dur)
                metrics = create_metrics(found, extreme, dur, bpm)
                # plot_data(data, filtered, found['index'], file)
                write_json(file, metrics)
                export_excel.append(metrics['num_beats'])
                numb = file.split('.')[0]
                numb = numb.split('data')[1]
                file_number.append(numb)
        except IndexError:
            print('Ignore Folder')

    write_excel(file_number, export_excel, 'Beat_Tracking.xlsx')


if __name__ == "__main__":
    main()
